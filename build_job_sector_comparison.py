from pathlib import Path
import re
import math
import statistics
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart.label import DataLabelList

BASE = Path(r'C:\UofA\BAT_498\Job_Sectors_2014-2024')
OUTPUT_DIR = Path(r'C:\Users\stars\Documents\Codex\2026-04-30\files-mentioned-by-the-user-administration\outputs\job_sector_comparison')
OUTPUT = OUTPUT_DIR / 'Job_Sector_Comparison_Averages.xlsx'
FILES = [
    ('Administration', BASE / 'Administration_2014-2024.xlsx'),
    ('Construction', BASE / 'Construction_2014-2024.xlsx'),
    ('Management of Companies', BASE / 'Managment_of_Comapanies_2014-2024.xlsx'),
    ('Medical Equipment', BASE / 'Medical_Equipment_2014-2024.xlsx'),
    ('School', BASE / 'School_2014-2024.xlsx'),
    ('Software Publishers', BASE / 'Software_Publishers_2014-2024.xlsx'),
]
EXCLUDE_COLUMNS = {'industry_code', 'year'}
TEXT_COLUMNS = {'industry_title', 'area_title', 'own_title'}
PREFERRED_METRICS = [
    'annual_avg_estabs_count',
    'annual_avg_emplvl',
    'total_annual_wages',
    'avg_annual_pay',
    'oty_total_annual_wages_chg',
    'oty_total_annual_wages_pct_chg',
    'oty_avg_annual_pay_chg',
    'oty_avg_annual_pay_pct_chg',
    'oty_annual_avg_emplvl_chg',
    'oty_annual_avg_emplvl_pct_chg',
]


def year_from_sheet(name):
    m = re.search(r'(20\d{2})', str(name))
    return int(m.group(1)) if m else None


def clean_sheet_name(name):
    cleaned = re.sub(r'[:\\/?*\[\]]', ' ', name)[:31].strip()
    return cleaned or 'Sheet'


def number_format_for(metric):
    if 'pct' in metric:
        return '0.00'
    if 'wages' in metric or 'pay' in metric:
        return '$#,##0.00'
    return '#,##0.00'


def pretty_metric(metric):
    return metric.replace('_', ' ').title()


def append_table(ws, name, first_row, first_col, last_row, last_col, style='TableStyleMedium2'):
    ref = f'{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}'
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)


def style_header(ws, row, start_col, end_col, fill='1F4E79'):
    for cell in ws[row][start_col-1:end_col]:
        cell.fill = PatternFill('solid', fgColor=fill)
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(bottom=Side(style='thin', color='D9EAF7'))


def autosize(ws, max_width=42):
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 10
        for cell in ws[letter]:
            value = cell.value
            if value is None:
                continue
            width = max(width, min(max_width, len(str(value)) + 2))
        ws.column_dimensions[letter].width = width


def add_title(ws, title, subtitle=None):
    ws['A1'] = title
    ws['A1'].font = Font(size=18, bold=True, color='17365D')
    if subtitle:
        ws['A2'] = subtitle
        ws['A2'].font = Font(italic=True, color='555555')
        ws['A2'].alignment = Alignment(wrap_text=True)


records = []
source_rows = []
metric_seen = []
metric_set = set()
sector_year_values = defaultdict(list)

for sector, file_path in FILES:
    xls = pd.ExcelFile(file_path)
    for sheet_name in xls.sheet_names:
        year = year_from_sheet(sheet_name)
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df.columns = [str(c).strip() if str(c).strip() else f'Column {i+1}' for i, c in enumerate(df.columns)]
        rows = len(df)
        numeric_candidates = []
        for column in df.columns:
            normalized = column.strip().lower()
            if normalized in EXCLUDE_COLUMNS or normalized in TEXT_COLUMNS:
                continue
            series = pd.to_numeric(df[column], errors='coerce')
            count = int(series.notna().sum())
            if count > 0:
                average = float(series.mean())
                numeric_candidates.append(column)
                records.append({
                    'Sector': sector,
                    'Year': year,
                    'Sheet': sheet_name,
                    'Metric': column,
                    'Metric Label': pretty_metric(column),
                    'Average': average,
                    'Numeric Count': count,
                    'Rows': rows,
                    'Source File': str(file_path),
                })
                if column not in metric_set:
                    metric_set.add(column)
                    metric_seen.append(column)
                sector_year_values[(sector, year)].append(average)
        source_rows.append({
            'Sector': sector,
            'Workbook': file_path.name,
            'Sheet': sheet_name,
            'Year': year,
            'Rows Read': rows,
            'Numeric Metrics Averaged': len(numeric_candidates),
            'Excluded Identifier/Text Columns': ', '.join(sorted(EXCLUDE_COLUMNS | TEXT_COLUMNS)),
        })

records_df = pd.DataFrame(records)
source_df = pd.DataFrame(source_rows)
metric_order = [m for m in PREFERRED_METRICS if m in metric_set] + [m for m in metric_seen if m not in PREFERRED_METRICS]
sectors = [s for s, _ in FILES]
years = sorted(y for y in records_df['Year'].dropna().unique())

# Tables for workbook
long_rows = records_df[['Sector', 'Year', 'Metric', 'Metric Label', 'Average', 'Numeric Count', 'Rows', 'Sheet', 'Source File']].sort_values(['Metric', 'Sector', 'Year']).values.tolist()
sector_year_summary = []
for sector in sectors:
    for year in years:
        vals = sector_year_values.get((sector, year), [])
        sector_year_summary.append([sector, int(year), statistics.mean(vals) if vals else None, len(vals)])

metric_by_sector_rows = []
for metric in metric_order:
    row = [pretty_metric(metric)]
    for sector in sectors:
        vals = records_df[(records_df['Sector'] == sector) & (records_df['Metric'] == metric)]['Average'].dropna().tolist()
        row.append(statistics.mean(vals) if vals else None)
    metric_by_sector_rows.append(row)

wb = Workbook()
ws_dash = wb.active
ws_dash.title = 'Dashboard'
ws_long = wb.create_sheet('All Averages')
ws_metric = wb.create_sheet('Metric by Sector')
ws_year = wb.create_sheet('Sector-Year Summary')
ws_source = wb.create_sheet('Source Detail')

for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False

# Dashboard
add_title(ws_dash, 'Job Sector Comparison: Column Averages', 'Numeric metrics are averaged within each yearly tab for each sector workbook. Identifier/text fields such as industry_code and year are excluded from metric averages.')
ws_dash.append([])
ws_dash.append(['Sector', '2014-2024 Overall Average Across Metrics', 'Metric-Year Averages Included'])
dash_header_row = ws_dash.max_row
for sector in sectors:
    vals = records_df[records_df['Sector'] == sector]['Average'].dropna().tolist()
    ws_dash.append([sector, statistics.mean(vals) if vals else None, len(vals)])
style_header(ws_dash, dash_header_row, 1, 3)
append_table(ws_dash, 'DashboardSectorTable', dash_header_row, 1, ws_dash.max_row, 3)
for row in range(dash_header_row + 1, ws_dash.max_row + 1):
    ws_dash.cell(row, 2).number_format = '#,##0.00'

chart = BarChart()
chart.title = 'Overall Average Across Metrics by Sector'
chart.y_axis.title = 'Average'
chart.x_axis.title = 'Sector'
chart.add_data(Reference(ws_dash, min_col=2, min_row=dash_header_row, max_row=ws_dash.max_row), titles_from_data=True)
chart.set_categories(Reference(ws_dash, min_col=1, min_row=dash_header_row + 1, max_row=ws_dash.max_row))
chart.height = 9
chart.width = 18
chart.legend = None
ws_dash.add_chart(chart, 'E4')

ws_dash['A13'] = 'Important note'
ws_dash['A13'].font = Font(bold=True, color='17365D')
ws_dash['A14'] = 'The workbook also includes graph-ready long data, a metric-by-sector comparison table, sector-year summaries, and one chart sheet per sector.'
ws_dash['A14'].alignment = Alignment(wrap_text=True)

# All Averages
all_headers = ['Sector', 'Year', 'Metric', 'Metric Label', 'Average', 'Numeric Count', 'Rows', 'Sheet', 'Source File']
ws_long.append(all_headers)
for row in long_rows:
    ws_long.append(row)
style_header(ws_long, 1, 1, len(all_headers))
append_table(ws_long, 'AllAveragesTable', 1, 1, ws_long.max_row, len(all_headers))
for row in range(2, ws_long.max_row + 1):
    ws_long.cell(row, 5).number_format = '#,##0.00'
ws_long.freeze_panes = 'A2'

# Metric by Sector
ws_metric.append(['Metric', *sectors])
for row in metric_by_sector_rows:
    ws_metric.append(row)
style_header(ws_metric, 1, 1, 1 + len(sectors), fill='7030A0')
append_table(ws_metric, 'MetricBySectorTable', 1, 1, ws_metric.max_row, ws_metric.max_column, style='TableStyleMedium5')
for row in range(2, ws_metric.max_row + 1):
    for col in range(2, ws_metric.max_column + 1):
        metric_name = metric_order[row - 2]
        ws_metric.cell(row, col).number_format = number_format_for(metric_name)
ws_metric.freeze_panes = 'B2'
if ws_metric.max_row > 2:
    ws_metric.conditional_formatting.add(f'B2:{get_column_letter(ws_metric.max_column)}{ws_metric.max_row}', ColorScaleRule(start_type='min', start_color='F8696B', mid_type='percentile', mid_value=50, mid_color='FFEB84', end_type='max', end_color='63BE7B'))

metric_chart = BarChart()
metric_chart.title = 'Metric Averages by Sector'
metric_chart.y_axis.title = 'Average'
metric_chart.x_axis.title = 'Metric'
metric_chart.add_data(Reference(ws_metric, min_col=2, max_col=ws_metric.max_column, min_row=1, max_row=min(ws_metric.max_row, 11)), titles_from_data=True)
metric_chart.set_categories(Reference(ws_metric, min_col=1, min_row=2, max_row=min(ws_metric.max_row, 11)))
metric_chart.height = 10
metric_chart.width = 22
ws_metric.add_chart(metric_chart, 'I2')

# Sector-Year Summary
ws_year.append(['Sector', 'Year', 'Average Across Metrics', 'Metrics Included'])
for row in sector_year_summary:
    ws_year.append(row)
style_header(ws_year, 1, 1, 4, fill='385723')
append_table(ws_year, 'SectorYearTable', 1, 1, ws_year.max_row, 4, style='TableStyleMedium4')
for row in range(2, ws_year.max_row + 1):
    ws_year.cell(row, 3).number_format = '#,##0.00'
ws_year.freeze_panes = 'A2'

# Helper area for line chart on year summary
start_col = 7
ws_year.cell(1, start_col, 'Year')
for i, sector in enumerate(sectors, start_col + 1):
    ws_year.cell(1, i, sector)
for r, year in enumerate(years, 2):
    ws_year.cell(r, start_col, int(year))
    for c, sector in enumerate(sectors, start_col + 1):
        vals = sector_year_values.get((sector, year), [])
        ws_year.cell(r, c, statistics.mean(vals) if vals else None)
        ws_year.cell(r, c).number_format = '#,##0.00'
style_header(ws_year, 1, start_col, start_col + len(sectors), fill='5B9BD5')
line = LineChart()
line.title = 'Average Across Metrics by Year'
line.y_axis.title = 'Average'
line.x_axis.title = 'Year'
line.add_data(Reference(ws_year, min_col=start_col + 1, max_col=start_col + len(sectors), min_row=1, max_row=1 + len(years)), titles_from_data=True)
line.set_categories(Reference(ws_year, min_col=start_col, min_row=2, max_row=1 + len(years)))
line.height = 11
line.width = 24
ws_year.add_chart(line, 'O2')

# Source Detail
ws_source.append(list(source_df.columns))
for row in source_df.values.tolist():
    ws_source.append(row)
style_header(ws_source, 1, 1, ws_source.max_column, fill='666666')
append_table(ws_source, 'SourceDetailTable', 1, 1, ws_source.max_row, ws_source.max_column, style='TableStyleMedium7')
ws_source.freeze_panes = 'A2'

# One sheet per sector with yearly metric averages and chart
for sector in sectors:
    sheet_name = clean_sheet_name(sector)
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    add_title(ws, f'{sector}: Yearly Column Averages')
    ws.append([])
    ws.append(['Year', *[pretty_metric(m) for m in metric_order]])
    header_row = ws.max_row
    for year in years:
        row = [int(year)]
        for metric in metric_order:
            vals = records_df[(records_df['Sector'] == sector) & (records_df['Year'] == year) & (records_df['Metric'] == metric)]['Average'].dropna().tolist()
            row.append(vals[0] if vals else None)
        ws.append(row)
    style_header(ws, header_row, 1, 1 + len(metric_order))
    append_table(ws, re.sub(r'\W+', '', sector)[:20] + 'YearlyTable', header_row, 1, ws.max_row, ws.max_column)
    for row in range(header_row + 1, ws.max_row + 1):
        for col in range(2, ws.max_column + 1):
            ws.cell(row, col).number_format = number_format_for(metric_order[col - 2])
    if metric_order:
        line = LineChart()
        line.title = f'{sector}: Selected Metrics Over Time'
        line.y_axis.title = 'Average'
        line.x_axis.title = 'Year'
        max_metric_cols = min(4, len(metric_order))
        line.add_data(Reference(ws, min_col=2, max_col=1 + max_metric_cols, min_row=header_row, max_row=ws.max_row), titles_from_data=True)
        line.set_categories(Reference(ws, min_col=1, min_row=header_row + 1, max_row=ws.max_row))
        line.height = 10
        line.width = 22
        ws.add_chart(line, 'O4')
    ws.freeze_panes = 'B5'
    autosize(ws)

for ws in wb.worksheets:
    autosize(ws)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=cell.alignment.wrap_text)

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
wb.save(OUTPUT)

# Validate the workbook can reopen and charts/sheets persisted.
check = load_workbook(OUTPUT, read_only=False, data_only=False)
print(f'Saved: {OUTPUT}')
print(f'Sheets: {len(check.sheetnames)}')
print(f'Records: {len(records_df)}')
print(f'Metrics: {len(metric_order)}')
print(f'Sectors: {len(sectors)}; Years: {len(years)}')

